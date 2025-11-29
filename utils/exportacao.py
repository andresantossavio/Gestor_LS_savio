"""
Utilitários para exportação de dados em Excel e PDF.
"""
from typing import List
from datetime import date
import io


def exportar_tarefas_excel(tarefas: List, filepath: str = None) -> bytes:
    """
    Exporta lista de tarefas para Excel.
    
    Args:
        tarefas: Lista de objetos Tarefa
        filepath: Caminho do arquivo (opcional, se None retorna bytes)
    
    Returns:
        Bytes do arquivo Excel
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise ImportError("openpyxl não está instalado. Execute: pip install openpyxl")
    
    # Cria workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tarefas"
    
    # Cabeçalho
    headers = [
        "ID", "Tipo", "Processo", "Cliente", "Responsável",
        "Prazo Administrativo", "Prazo Fatal", "Status", "Etapa Workflow",
        "Município", "UF", "Criado Em"
    ]
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Dados
    for row_idx, tarefa in enumerate(tarefas, start=2):
        ws.cell(row=row_idx, column=1, value=tarefa.id)
        ws.cell(row=row_idx, column=2, value=tarefa.tipo_tarefa.nome if tarefa.tipo_tarefa else "")
        ws.cell(row=row_idx, column=3, value=tarefa.processo.numero if tarefa.processo else "")
        ws.cell(row=row_idx, column=4, value=tarefa.processo.cliente.nome if tarefa.processo and tarefa.processo.cliente else "")
        ws.cell(row=row_idx, column=5, value=tarefa.responsavel.nome if tarefa.responsavel else "")
        ws.cell(row=row_idx, column=6, value=tarefa.prazo_administrativo.strftime("%d/%m/%Y") if tarefa.prazo_administrativo else "")
        ws.cell(row=row_idx, column=7, value=tarefa.prazo_fatal.strftime("%d/%m/%Y") if tarefa.prazo_fatal else "")
        ws.cell(row=row_idx, column=8, value=tarefa.status)
        ws.cell(row=row_idx, column=9, value=tarefa.etapa_workflow_atual)
        
        if tarefa.processo and tarefa.processo.municipio:
            ws.cell(row=row_idx, column=10, value=tarefa.processo.municipio.nome)
            ws.cell(row=row_idx, column=11, value=tarefa.processo.municipio.uf)
        
        ws.cell(row=row_idx, column=12, value=tarefa.criado_em.strftime("%d/%m/%Y %H:%M") if tarefa.criado_em else "")
    
    # Ajusta largura das colunas
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    # Salva ou retorna bytes
    if filepath:
        wb.save(filepath)
        return None
    else:
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()


def exportar_tarefas_pdf(tarefas: List, filepath: str = None) -> bytes:
    """
    Exporta lista de tarefas para PDF.
    
    Args:
        tarefas: Lista de objetos Tarefa
        filepath: Caminho do arquivo (opcional, se None retorna bytes)
    
    Returns:
        Bytes do arquivo PDF
    """
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
    except ImportError:
        raise ImportError("reportlab não está instalado. Execute: pip install reportlab")
    
    # Cria documento
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer if not filepath else filepath,
        pagesize=landscape(A4),
        rightMargin=1*cm,
        leftMargin=1*cm,
        topMargin=1*cm,
        bottomMargin=1*cm
    )
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Título
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#366092'),
        spaceAfter=30,
        alignment=1  # Center
    )
    elements.append(Paragraph("Relatório de Tarefas", title_style))
    elements.append(Spacer(1, 0.5*cm))
    
    # Tabela de dados
    data = [["ID", "Tipo", "Processo", "Cliente", "Responsável", "Prazo Fatal", "Status"]]
    
    for tarefa in tarefas:
        data.append([
            str(tarefa.id),
            tarefa.tipo_tarefa.nome[:30] if tarefa.tipo_tarefa else "",
            tarefa.processo.numero[:20] if tarefa.processo else "",
            tarefa.processo.cliente.nome[:25] if tarefa.processo and tarefa.processo.cliente else "",
            tarefa.responsavel.nome[:20] if tarefa.responsavel else "",
            tarefa.prazo_fatal.strftime("%d/%m/%Y") if tarefa.prazo_fatal else "",
            tarefa.status
        ])
    
    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
    ]))
    
    elements.append(table)
    
    # Rodapé
    elements.append(Spacer(1, 1*cm))
    footer_style = ParagraphStyle(
        'Footer',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.grey
    )
    elements.append(Paragraph(
        f"Gerado em: {date.today().strftime('%d/%m/%Y')} | Total de tarefas: {len(tarefas)}",
        footer_style
    ))
    
    # Gera PDF
    doc.build(elements)
    
    if filepath:
        return None
    else:
        buffer.seek(0)
        return buffer.getvalue()


def exportar_dashboard_excel(estatisticas: dict, filepath: str = None) -> bytes:
    """
    Exporta estatísticas do dashboard para Excel.
    
    Args:
        estatisticas: Dicionário com dados do dashboard
        filepath: Caminho do arquivo (opcional)
    
    Returns:
        Bytes do arquivo Excel ou None se filepath fornecido
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.chart import BarChart, PieChart, Reference
    except ImportError:
        raise ImportError("openpyxl não está instalado")
    
    wb = openpyxl.Workbook()
    
    # Remove sheet padrão e cria sheets customizadas
    wb.remove(wb.active)
    
    # Sheet 1: Resumo
    ws_resumo = wb.create_sheet("Resumo")
    ws_resumo.append(["Métrica", "Valor"])
    
    for key, value in estatisticas.items():
        if isinstance(value, (int, float, str)):
            ws_resumo.append([key, value])
    
    # Aplica estilo ao cabeçalho
    for cell in ws_resumo[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
    
    # Salva ou retorna
    if filepath:
        wb.save(filepath)
        return None
    else:
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
